"""Descarga incrementalmente las comunicaciones del BCRA desde 2026."""

from __future__ import annotations

from concurrent.futures import ThreadPoolExecutor, as_completed
from copy import copy
from datetime import date, datetime
from pathlib import Path
import re
from threading import local
from urllib.parse import urljoin, urlparse

import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from pypdf import PdfReader
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry


RAIZ_PROYECTO = Path(__file__).resolve().parents[1]
ARCHIVO_BD = RAIZ_PROYECTO / "BD.xlsx"
CARPETA_COMUNICACIONES = RAIZ_PROYECTO / "fuentes_BD" / "BCRA" / "Comunicaciones"
API_URL = "https://www.bcra.gob.ar/api/endpoints/buscador-comunicaciones.php"
BASE_URL = "https://www.bcra.gob.ar"

# ========================= CONFIGURACIÓN DE LA BÚSQUEDA =========================
# Fecha inicial INCLUSIVE. Para traer todo el histórico del BCRA usar date(2006, 1, 1).
FECHA_DESDE = date(2026, 1, 1)
# Fecha final INCLUSIVE. Se actualiza automáticamente al día en que se ejecuta el scraper.
# Para fijar una fecha manual, reemplazar date.today() por date(AAAA, MM, DD).
FECHA_HASTA = date.today()
# Tipos a buscar: usar "ALL" para A, B, C y P; "NONE" solo para registros sin tipo.
# Ejemplos: ("A", "B") | ("ALL",) | ("NONE",)
TIPOS_BUSQUEDA = ("ALL",)
# Circulares asociadas: usar "ALL" para incluir con y sin circular; "NONE" solo sin circular.
# Ejemplos: ("CAMEX", "OPASI") | ("ALL",) | ("NONE",)
CIRCULARES_BUSQUEDA = ("ALL",)
# ================================================================================

TIPOS_DISPONIBLES = ("A", "B", "C", "P")
HOJA_BD = "Comunicaciones BCRA"
MAX_DESCARGAS_SIMULTANEAS = 4

# Agrupamientos publicados por el BCRA, más SINAP, presente en referencias recientes.
CIRCULARES_CONOCIDAS = (
    "CAMCO", "CAMEX", "CIRMO", "CONAU", "COPEX", "CREFI", "LISOL", "OPASI",
    "OPRAC", "REFEX", "REMON", "RUNOR", "SEPEX", "SERVI", "SINAP", "TINAC",
)

COLUMNAS = [
    "Fecha emisión",
    "Tipo",
    "Número",
    "Tipo y número",
    "Circular asociada",
    "Referencia",
    "Boletín oficial",
    "Fecha publicación",
    "URL",
    "Texto local",
]

_SESIONES_HILO = local()


def crear_sesion() -> requests.Session:
    """Crea una sesión con reintentos para las respuestas inestables del sitio."""
    reintentos = Retry(
        total=5,
        connect=5,
        read=5,
        backoff_factor=1,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=frozenset({"GET", "POST"}),
    )
    sesion = requests.Session()
    sesion.headers.update({"User-Agent": "EcoSeriesScraper/1.0 (datos públicos BCRA)"})
    sesion.mount("https://", HTTPAdapter(max_retries=reintentos))
    return sesion


def _sesion_hilo() -> requests.Session:
    sesion = getattr(_SESIONES_HILO, "sesion", None)
    if sesion is None:
        sesion = crear_sesion()
        _SESIONES_HILO.sesion = sesion
    return sesion


def _fecha(valor: object) -> date | None:
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    texto = str(valor or "").strip()
    if not texto:
        return None
    try:
        return datetime.strptime(texto, "%Y-%m-%d").date()
    except ValueError:
        return None


def _url_registro(registro: dict[str, object]) -> str:
    enlace = str(registro.get("link_url") or registro.get("pdf_path") or "").strip()
    return urljoin(BASE_URL, enlace) if enlace else ""


def _circulares_en_referencia(titulo: object) -> set[str]:
    texto = str(titulo or "").upper()
    encontradas = {
        circular
        for circular in CIRCULARES_CONOCIDAS
        if re.search(rf"(?<![A-Z]){re.escape(circular)}(?![A-Z])", texto)
    }
    # Captura agrupamientos nuevos del BCRA que todavía no estén en el catálogo.
    encontradas.update(re.findall(r"(?<![A-Z])([A-Z]{4,5})(?=\s*[-–]?\s*\d)", texto))
    return encontradas


def _seleccionado(valor: str, configuracion: tuple[str, ...], todos: str = "ALL") -> bool:
    """Evalúa una selección ALL/NONE o una lista explícita de valores."""
    valores = {str(item).strip().upper() for item in configuracion}
    if todos in valores:
        return True
    if "NONE" in valores:
        return not valor
    return valor.upper() in valores


def _tipo_seleccionado(tipo: str) -> bool:
    return _seleccionado(tipo.strip().upper(), TIPOS_BUSQUEDA)


def _circular_seleccionada(circulares: set[str]) -> bool:
    configuracion = {str(item).strip().upper() for item in CIRCULARES_BUSQUEDA}
    if "ALL" in configuracion:
        return True
    if "NONE" in configuracion:
        return not circulares
    return bool(circulares & configuracion)


def consultar_tipo_fecha(
    sesion: requests.Session,
    tipo: str | None,
    fecha_desde: date,
    fecha_hasta: date,
) -> list[dict[str, object]]:
    """Obtiene comunicaciones por tipo o por fecha general si tipo es None."""
    pagina = 1
    seleccionados: list[dict[str, object]] = []
    while True:
        parametros = {
            "mode": "tipo-fecha" if tipo else "fecha",
            "fecha_desde": fecha_desde.isoformat(),
            "fecha_hasta": fecha_hasta.isoformat(),
            "paginaabsoluta": pagina,
            "tamanopagina": 100,
            "lang": "es",
        }
        respuesta = sesion.post(API_URL, data=parametros, timeout=(20, 120))
        respuesta.raise_for_status()
        contenido = respuesta.json()
        if not contenido.get("success"):
            raise RuntimeError(contenido.get("error") or f"Respuesta inválida para tipo {tipo or 'todos'}")

        datos = contenido.get("data") or {}
        registros = datos.get("registros") or []
        for original in registros:
            fecha_emision = _fecha(original.get("fecha_emision"))
            if fecha_emision is None or not fecha_desde <= fecha_emision <= fecha_hasta:
                continue
            registro = dict(original)
            registro["tipo"] = str(registro.get("tipo") or "").strip().upper()
            registro["circulares"] = _circulares_en_referencia(registro.get("titulo"))
            registro["url"] = _url_registro(registro)
            if _tipo_seleccionado(registro["tipo"]) and _circular_seleccionada(registro["circulares"]):
                seleccionados.append(registro)

        paginacion = datos.get("pagination") or {}
        total_paginas = int(paginacion.get("totalPages") or 1)
        if pagina >= total_paginas or not registros:
            break
        pagina += 1
    return seleccionados


def consolidar_registros(registros: list[dict[str, object]]) -> list[dict[str, object]]:
    """Deduplica por tipo y número y conserva todas las asociaciones temáticas."""
    unicos: dict[str, dict[str, object]] = {}
    for registro in registros:
        tipo = str(registro.get("tipo") or "").strip().upper()
        numero = str(registro.get("numero_formateado") or "").strip()
        clave = f"{tipo}{numero}"
        if not numero:
            continue
        if clave not in unicos:
            unicos[clave] = dict(registro)
            unicos[clave]["circulares"] = set(registro.get("circulares") or [])
        else:
            unicos[clave]["circulares"].update(registro.get("circulares") or [])
    resultado = list(unicos.values())
    resultado.sort(
        key=lambda x: (
            _fecha(x.get("fecha_emision")) or date.min,
            str(x.get("tipo")),
            int(x.get("numero_formateado") or 0),
        ),
        reverse=True,
    )
    return resultado


def _clave_registro(registro: dict[str, object]) -> str:
    tipo = str(registro.get("tipo") or "SIN_TIPO").strip().upper()
    return f"{tipo}{registro.get('numero_formateado', '')}"


def _nombre_archivo(registro: dict[str, object]) -> str:
    circulares = sorted(set(registro.get("circulares") or []))
    agrupamiento = " + ".join(circulares) if circulares else "SIN CIRCULAR"
    numero = str(registro.get("numero_formateado") or "").strip()
    return f"{agrupamiento} - {numero}"


def _texto_valido(ruta: Path) -> bool:
    if not ruta.is_file() or ruta.suffix.lower() != ".txt" or ruta.stat().st_size == 0:
        return False
    try:
        return bool(ruta.read_text(encoding="utf-8").strip())
    except UnicodeDecodeError:
        return False


def _pdf_valido(ruta: Path) -> bool:
    if not ruta.is_file() or ruta.stat().st_size == 0:
        return False
    with ruta.open("rb") as archivo:
        return archivo.read(5) == b"%PDF-"


def _extraer_texto_pdf(origen: Path, destino: Path) -> None:
    """Extrae texto y publica el TXT atómicamente; no acepta resultados vacíos."""
    lector = PdfReader(origen)
    paginas = [pagina.extract_text() or "" for pagina in lector.pages]
    texto = "\n\n".join(pagina.strip() for pagina in paginas if pagina.strip()).strip()
    if not texto:
        raise ValueError("El PDF no contiene texto extraíble; se conserva el PDF")
    temporal = destino.with_name(f"{destino.stem}.descarga{destino.suffix}")
    try:
        temporal.write_text(texto + "\n", encoding="utf-8", newline="\n")
        if not _texto_valido(temporal):
            raise ValueError("No se pudo validar el texto extraído")
        temporal.replace(destino)
    finally:
        temporal.unlink(missing_ok=True)


def _ruta_desde_bd(valor: object) -> Path | None:
    texto = str(valor or "").strip()
    if not texto:
        return None
    ruta = Path(texto.replace("\\", "/"))
    return ruta if ruta.is_absolute() else RAIZ_PROYECTO / ruta


def _destino_registro(registro: dict[str, object]) -> Path | None:
    url = str(registro.get("url") or "").strip()
    if not url:
        return None
    tipo = str(registro.get("tipo") or "SIN_TIPO").strip().upper()
    return CARPETA_COMUNICACIONES / tipo / f"{_nombre_archivo(registro)}.txt"


def _candidatos_existentes(
    registro: dict[str, object], fila_existente: list[object] | None
) -> list[Path]:
    tipo = str(registro.get("tipo") or "SIN_TIPO").strip().upper()
    numero = str(registro.get("numero_formateado") or "").strip()
    candidatos: list[Path] = []
    if fila_existente:
        ruta_bd = _ruta_desde_bd(fila_existente[9])
        if ruta_bd:
            candidatos.append(ruta_bd)
    carpeta_tipo = CARPETA_COMUNICACIONES / tipo
    if carpeta_tipo.exists():
        candidatos.extend(carpeta_tipo.glob(f"* - {numero}.*"))
    candidatos.extend(CARPETA_COMUNICACIONES.glob(f"{tipo}{numero}.*"))
    return candidatos


def _reutilizar_archivo(
    registro: dict[str, object], fila_existente: list[object] | None
) -> tuple[Path | None, str]:
    destino = _destino_registro(registro)
    if destino is None:
        return None, "sin_url"
    destino.parent.mkdir(parents=True, exist_ok=True)
    if _texto_valido(destino):
        return destino, "existente"
    for candidato in _candidatos_existentes(registro, fila_existente):
        if candidato == destino:
            continue
        if _texto_valido(candidato):
            candidato.replace(destino)
            return destino, "reorganizado"
        if candidato.suffix.lower() == ".pdf" and _pdf_valido(candidato):
            try:
                _extraer_texto_pdf(candidato, destino)
            except ValueError:
                return candidato, "pdf_sin_texto"
            candidato.unlink()
            return destino, "convertido"
    return None, "faltante"


def descargar_documento(
    registro: dict[str, object], fila_existente: list[object] | None = None
) -> tuple[Path | None, str]:
    """Reutiliza un archivo previo o descarga únicamente el que falta."""
    existente, estado = _reutilizar_archivo(registro, fila_existente)
    if existente is not None or estado == "sin_url":
        return existente, estado

    destino = _destino_registro(registro)
    assert destino is not None
    extension_fuente = Path(urlparse(str(registro["url"])).path).suffix.lower()
    temporal = destino.with_name(f"{destino.stem}.descarga{extension_fuente or '.html'}")
    try:
        with _sesion_hilo().get(str(registro["url"]), stream=True, timeout=(20, 180)) as respuesta:
            respuesta.raise_for_status()
            with temporal.open("wb") as archivo:
                for bloque in respuesta.iter_content(chunk_size=64 * 1024):
                    if bloque:
                        archivo.write(bloque)
        if _pdf_valido(temporal):
            try:
                _extraer_texto_pdf(temporal, destino)
            except ValueError:
                pdf_conservado = destino.with_suffix(".pdf")
                temporal.replace(pdf_conservado)
                return pdf_conservado, "pdf_sin_texto"
        else:
            texto = temporal.read_text(encoding="utf-8", errors="replace").strip()
            if not texto:
                raise ValueError("La comunicación descargada no contiene texto")
            destino.write_text(texto + "\n", encoding="utf-8", newline="\n")
        if not _texto_valido(destino):
            raise ValueError("No se pudo validar el texto de la comunicación")
        return destino, "descargado"
    finally:
        temporal.unlink(missing_ok=True)


def _fila_excel(registro: dict[str, object], archivo: Path | None) -> list[object]:
    fecha_emision = _fecha(registro.get("fecha_emision"))
    fecha_publicacion = _fecha(registro.get("fecha_boletin"))
    tipo = str(registro.get("tipo") or "").strip().upper()
    numero = int(registro.get("numero_formateado") or 0)
    circulares = ", ".join(sorted(set(registro.get("circulares") or [])))
    ruta_local = str(archivo.relative_to(RAIZ_PROYECTO)) if archivo else ""
    return [
        fecha_emision,
        tipo,
        numero,
        f"{tipo}{numero}" if tipo else f"SIN TIPO - {numero}",
        circulares,
        str(registro.get("titulo") or "").strip(),
        str(registro.get("nro_boletin") or "").strip(),
        fecha_publicacion,
        str(registro.get("url") or ""),
        ruta_local,
    ]


def cargar_registros_bd() -> dict[str, list[object]]:
    """Carga las rutas ya registradas para evitar volver a descargar archivos."""
    libro = load_workbook(ARCHIVO_BD, data_only=False, read_only=True)
    existentes: dict[str, list[object]] = {}
    if HOJA_BD in libro.sheetnames:
        hoja = libro[HOJA_BD]
        for valores in hoja.iter_rows(min_row=2, max_col=len(COLUMNAS), values_only=True):
            clave = str(valores[3] or "").strip()
            if clave:
                existentes[clave] = list(valores)
    libro.close()
    return existentes


def reorganizar_archivos_previos(existentes: dict[str, list[object]]) -> dict[str, int]:
    """Convierte la muestra anterior y la mueve a carpetas por tipo."""
    conteos = {"reorganizado": 0, "convertido": 0, "pdf_sin_texto": 0}
    for clave, fila in existentes.items():
        tipo = str(fila[1] or "SIN_TIPO").strip().upper()
        numero = int(fila[2] or clave[1:])
        circulares = {x.strip() for x in str(fila[4] or "").split(",") if x.strip()}
        registro = {
            "tipo": tipo,
            "numero_formateado": numero,
            "circulares": circulares,
            "url": str(fila[8] or ""),
        }
        archivo, estado = _reutilizar_archivo(registro, fila)
        if archivo:
            fila[9] = str(archivo.relative_to(RAIZ_PROYECTO))
        if estado in conteos:
            conteos[estado] += 1
    return conteos


def actualizar_hoja_bd(filas: list[list[object]]) -> int:
    """Reemplaza el inventario del período preservando estilos y las demás hojas."""
    filas = sorted(filas, key=lambda x: (x[0] or date.min, str(x[3])), reverse=True)
    libro = load_workbook(ARCHIVO_BD)
    hoja = libro[HOJA_BD] if HOJA_BD in libro.sheetnames else libro.create_sheet(HOJA_BD)
    filas_anteriores = hoja.max_row
    columnas_anteriores = hoja.max_column

    for columna, encabezado in enumerate(COLUMNAS, 1):
        celda = hoja.cell(1, columna)
        celda.value = encabezado
        if filas_anteriores == 1 and columnas_anteriores == 1:
            celda.fill = PatternFill("solid", fgColor="1F4E78")
            celda.font = Font(color="FFFFFF", bold=True)
            celda.alignment = Alignment(horizontal="center")

    for indice_fila, valores in enumerate(filas, 2):
        if indice_fila > filas_anteriores and filas_anteriores >= 2:
            for columna in range(1, len(COLUMNAS) + 1):
                origen = hoja.cell(2, min(columna, columnas_anteriores))
                destino = hoja.cell(indice_fila, columna)
                if origen.has_style:
                    destino._style = copy(origen._style)
        for columna, valor in enumerate(valores, 1):
            hoja.cell(indice_fila, columna).value = valor
        for columna in (1, 8):
            if hoja.cell(indice_fila, columna).value:
                hoja.cell(indice_fila, columna).number_format = "yyyy-mm-dd"
        url = str(hoja.cell(indice_fila, 9).value or "")
        if url:
            for columna in (4, 9):
                hoja.cell(indice_fila, columna).hyperlink = url
                hoja.cell(indice_fila, columna).style = "Hyperlink"

    ultima_fila = len(filas) + 1
    for fila in range(ultima_fila + 1, filas_anteriores + 1):
        for columna in range(1, max(columnas_anteriores, len(COLUMNAS)) + 1):
            hoja.cell(fila, columna).value = None

    if "TablaComunicacionesBCRA" in hoja.tables:
        del hoja.tables["TablaComunicacionesBCRA"]
    if filas:
        tabla = Table(displayName="TablaComunicacionesBCRA", ref=f"A1:J{ultima_fila}")
        tabla.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        hoja.add_table(tabla)
    hoja.freeze_panes = "A2"
    if filas_anteriores == 1 and columnas_anteriores == 1:
        for columna, ancho in enumerate([14, 8, 10, 16, 24, 70, 18, 18, 65, 55], 1):
            hoja.column_dimensions[get_column_letter(columna)].width = ancho
    libro.save(ARCHIVO_BD)
    libro.close()
    return len(filas)


def procesar(
    fecha_desde: date = FECHA_DESDE,
    fecha_hasta: date = FECHA_HASTA,
) -> dict[str, int]:
    """Actualiza el universo completo y descarga solo archivos nuevos o faltantes."""
    if fecha_desde > fecha_hasta:
        raise ValueError("fecha_desde no puede ser posterior a fecha_hasta")

    CARPETA_COMUNICACIONES.mkdir(parents=True, exist_ok=True)
    configuracion_tipos = {str(item).strip().upper() for item in TIPOS_BUSQUEDA}
    for tipo in (*TIPOS_DISPONIBLES, "SIN_TIPO"):
        (CARPETA_COMUNICACIONES / tipo).mkdir(parents=True, exist_ok=True)

    existentes = cargar_registros_bd()
    hay_archivos_heredados = any(
        ruta.is_file() for ruta in CARPETA_COMUNICACIONES.iterdir()
    )
    previos = (
        reorganizar_archivos_previos(existentes)
        if hay_archivos_heredados
        else {"reorganizado": 0, "convertido": 0, "pdf_sin_texto": 0}
    )

    encontrados: list[dict[str, object]] = []
    with crear_sesion() as sesion:
        if configuracion_tipos & {"ALL", "NONE"}:
            registros = consultar_tipo_fecha(sesion, None, fecha_desde, fecha_hasta)
            encontrados.extend(registros)
            print(f"BCRA búsqueda general: {len(registros)} comunicaciones", flush=True)
        else:
            for tipo in TIPOS_DISPONIBLES:
                if tipo not in configuracion_tipos:
                    continue
                registros = consultar_tipo_fecha(sesion, tipo, fecha_desde, fecha_hasta)
                encontrados.extend(registros)
                print(f"BCRA tipo {tipo}: {len(registros)} comunicaciones", flush=True)
    consolidados = consolidar_registros(encontrados)

    filas: list[list[object]] = []
    conteos = {
        "existente": 0,
        "reorganizado": 0,
        "convertido": 0,
        "descargado": 0,
        "pdf_sin_texto": 0,
        "sin_url": 0,
        "error": 0,
    }

    def tarea(registro: dict[str, object]):
        clave = _clave_registro(registro)
        try:
            archivo, estado = descargar_documento(registro, existentes.get(clave))
            return registro, archivo, estado, ""
        except (OSError, requests.RequestException, ValueError) as exc:
            return registro, None, "error", str(exc)

    with ThreadPoolExecutor(max_workers=MAX_DESCARGAS_SIMULTANEAS) as ejecutor:
        futuros = [ejecutor.submit(tarea, registro) for registro in consolidados]
        for indice, futuro in enumerate(as_completed(futuros), 1):
            registro, archivo, estado, error = futuro.result()
            conteos[estado] += 1
            filas.append(_fila_excel(registro, archivo))
            if error:
                print(f"Error {_clave_registro(registro)}: {error}", flush=True)
            if indice % 25 == 0 or indice == len(futuros):
                print(f"Archivos BCRA: {indice}/{len(futuros)} procesados", flush=True)

    total_bd = actualizar_hoja_bd(filas)
    # Mantiene la fecha de inicio de la serie administrativa alineada con la configuración.
    from tools.generar_codificacion import generar as generar_codificacion
    generar_codificacion()
    resumen = {
        "encontradas": len(consolidados),
        "ya_existentes": conteos["existente"],
        "reorganizadas": previos["reorganizado"] + conteos["reorganizado"],
        "pdfs_convertidos": previos["convertido"] + conteos["convertido"],
        "descargadas": conteos["descargado"],
        "pdfs_sin_texto": previos["pdf_sin_texto"] + conteos["pdf_sin_texto"],
        "sin_url": conteos["sin_url"],
        "errores": conteos["error"],
        "registros_bd": total_bd,
    }
    print(f"BCRA terminado: {resumen}", flush=True)
    return resumen


def ejecutar() -> None:
    procesar()


if __name__ == "__main__":
    ejecutar()
