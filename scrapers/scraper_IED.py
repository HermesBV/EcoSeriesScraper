"""Descarga y procesa las series del Informe Económico al Día (IED)."""

from __future__ import annotations

import re
import ssl
import warnings
from datetime import datetime
from pathlib import Path
from zipfile import BadZipFile, ZipFile

import certifi
import numpy as np
import pandas as pd
import requests
from openpyxl import load_workbook
from requests.adapters import HTTPAdapter
from tqdm import tqdm
from urllib3.util.retry import Retry

from scrapers.ied_inventory import discover_and_extract, load_existing_catalog


RAIZ_PROYECTO = Path(__file__).resolve().parents[1]
CARPETA_FUENTE = RAIZ_PROYECTO / "fuentes_BD" / "MECON" / "IED"
CARPETA_LOGS = RAIZ_PROYECTO / "logs"
ARCHIVO_BD = RAIZ_PROYECTO / "BD.xlsx"
HOJAS_ADMINISTRATIVAS = {
    "Referencias", "Introduccion_Codigos", "Referencia_Codigos",
    "Mapa_Tematico", "Parentesco_Codigos", "Codificacion",
}
HOJAS_AJENAS_IED = {"Comunicaciones BCRA"}

EXCEL_URLS = {
    "empleo_ingresos.xlsx": "https://www.economia.gob.ar/download/infoeco/apendice3a.xlsx",
    "sector_externo.xlsx": "https://www.economia.gob.ar/download/infoeco/apendice5.xlsx",
    "internacional.xlsx": "https://www.economia.gob.ar/download/infoeco/internacional_ied.xlsx",
    "dinero_bancos.xlsx": "https://www.economia.gob.ar/download/infoeco/apendice8.xlsx",
    "precios.xlsx": "https://www.economia.gob.ar/download/infoeco/apendice4.xlsx",
    "finanzas.xlsx": "https://www.economia.gob.ar/download/infoeco/apendice-financiero.xlsx",
    "finanzas_publicas.xlsx": "https://www.economia.gob.ar/download/infoeco/apendice6.xlsx",
    "actividad.xlsx": "https://www.economia.gob.ar/download/infoeco/actividad_ied.xlsx",
}

MAPEO_ORIGEN_ARCHIVO = {
    "Empleo e Ingresos: Apendice3A": "empleo_ingresos.xlsx",
    "Actividad: Actividad_IED": "actividad.xlsx",
    "Contexto Internacional": "internacional.xlsx",
    "Precios: Apendice4": "precios.xlsx",
    "Finanzas: Apendice-Financiero": "finanzas.xlsx",
    "Finanzas Públicas: Apendice6": "finanzas_publicas.xlsx",
    "Sector Externo: Apendice5": "sector_externo.xlsx",
    "Dinero y Bancos: Apendice8": "dinero_bancos.xlsx",
}

# Correcciones de nombres históricos cuya frecuencia no coincidía con los datos.
RENOMBRES_PESTANAS = {
    "EPH Puntual - Pel M": "EPH Puntual - Pel S",
    "IPC Mensual M": "IPC Trimestral T",
    "TCN D": "TCN A",
    "ITCRM D": "ITCRM A",
    "EXPO IMPO P Y Q T": "EXPO IMPO P Y Q A",
    "TDI T": "TDI A",
    "IPI T": "IPI A",
    "8.8. Tipos de cambio historicos": "8.8 Tipos de cambio D",
    "8.8. Tipos de cambio historicos D": "8.8 Tipos de cambio D",
    "8.3 Prest. de Entidades finac.": "8.3 Préstamos entidades M",
    "8.3 Prest. de Entidades finac. M": "8.3 Préstamos entidades M",
}

MESES = {
    "ene": "Jan", "enero": "Jan", "feb": "Feb", "febrero": "Feb",
    "mar": "Mar", "marzo": "Mar", "abr": "Apr", "abril": "Apr",
    "may": "May", "mayo": "May", "jun": "Jun", "junio": "Jun",
    "jul": "Jul", "julio": "Jul", "ago": "Aug", "agosto": "Aug",
    "sep": "Sep", "sept": "Sep", "septiembre": "Sep",
    "oct": "Oct", "octubre": "Oct", "nov": "Nov", "noviembre": "Nov",
    "dic": "Dec", "diciembre": "Dec",
}


def escribir_log(id_serie: str, estado: str, mensaje: str = "") -> None:
    CARPETA_LOGS.mkdir(parents=True, exist_ok=True)
    ahora = datetime.now()
    log_file = CARPETA_LOGS / f"log_{ahora:%Y%m%d}.txt"
    mensaje = mensaje.replace("\n", " || ")
    partes = [f"{ahora:%Y%m%d_%H%M%S}", id_serie, estado]
    if mensaje:
        partes.append(mensaje)
    with log_file.open("a", encoding="utf-8") as archivo:
        archivo.write("|".join(partes) + "\n")


def limpiar_nombres_definidos(ruta_archivo: Path) -> None:
    """Elimina rangos con nombre defectuosos de algunos libros del IED."""
    try:
        libro = load_workbook(ruta_archivo)
        for nombre in list(libro.defined_names):
            del libro.defined_names[nombre]
        libro.save(ruta_archivo)
    except Exception as exc:
        escribir_log("SISTEMA", "ADVERTENCIA_LIMPIEZA", f"{ruta_archivo.name}: {exc}")


def _excel_valido(ruta: Path) -> bool:
    """Comprueba que una descarga o copia local sea un libro Excel legible."""
    if not ruta.is_file() or ruta.stat().st_size == 0:
        return False
    try:
        with ZipFile(ruta) as archivo:
            nombres = set(archivo.namelist())
            if "[Content_Types].xml" not in nombres or "xl/workbook.xml" not in nombres:
                return False
            with archivo.open("xl/workbook.xml") as libro:
                return bool(libro.read(1))
    except (BadZipFile, OSError, KeyError):
        return False


class _AdaptadorSSL(HTTPAdapter):
    """Adaptador que incorpora el almacén de certificados del sistema."""

    def __init__(self, contexto_ssl: ssl.SSLContext, **opciones: object) -> None:
        self.contexto_ssl = contexto_ssl
        super().__init__(**opciones)

    def init_poolmanager(self, *args: object, **opciones: object) -> None:
        opciones["ssl_context"] = self.contexto_ssl
        super().init_poolmanager(*args, **opciones)


def _crear_sesion_descarga() -> requests.Session:
    """Crea una sesión robusta usando también los certificados del sistema."""
    contexto_ssl = ssl.create_default_context()
    contexto_ssl.load_verify_locations(cafile=certifi.where())
    reintentos = Retry(
        total=None,
        connect=3,
        read=3,
        status=3,
        other=0,
        backoff_factor=1,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=frozenset({"GET"}),
    )
    sesion = requests.Session()
    # Ignora HTTP(S)_PROXY del entorno: puede apuntar a un proxy local
    # inexistente aunque los enlaces funcionen desde el navegador.
    sesion.trust_env = False
    sesion.headers.update({"User-Agent": "EcoSeriesScraper/1.0 (datos públicos MECON)"})
    sesion.mount(
        "https://",
        _AdaptadorSSL(contexto_ssl, max_retries=reintentos),
    )
    return sesion


def descargar_excels() -> dict[str, Path]:
    """Descarga cada libro de forma atómica para no destruir una copia válida."""
    CARPETA_FUENTE.mkdir(parents=True, exist_ok=True)
    descargados: dict[str, Path] = {}
    with _crear_sesion_descarga() as sesion:
        for nombre, url in tqdm(EXCEL_URLS.items(), desc="Descargando IED"):
            destino = CARPETA_FUENTE / nombre
            temporal = destino.with_name(f"{destino.stem}.descarga.xlsx")
            try:
                try:
                    respuesta = sesion.get(url, stream=True, timeout=(15, 120))
                except requests.exceptions.SSLError:
                    # El servidor de Economia entrega una cadena que algunas
                    # instalaciones no pueden validar. Reintentamos solo este
                    # caso para no dejar de actualizar los libros diarios.
                    # La descarga sigue siendo HTTPS y el Excel se valida
                    # antes de reemplazar la copia local.
                    print(f"\nAdvertencia: certificado no verificable para {nombre}; se reintenta por HTTPS.")
                    with requests.Session() as sesion_sin_verificacion:
                        sesion_sin_verificacion.trust_env = False
                        sesion_sin_verificacion.headers.update(sesion.headers)
                        with warnings.catch_warnings():
                            warnings.simplefilter("ignore")
                            respuesta = sesion_sin_verificacion.get(
                                url, stream=True, timeout=(15, 120), verify=False
                            )
                        _guardar_descarga(respuesta, temporal)
                else:
                    _guardar_descarga(respuesta, temporal)
                if not _excel_valido(temporal):
                    raise ValueError("la respuesta descargada no es un archivo Excel válido")
                temporal.replace(destino)
                limpiar_nombres_definidos(destino)
                descargados[nombre] = destino
            except Exception as exc:
                temporal.unlink(missing_ok=True)
                if _excel_valido(destino):
                    descargados[nombre] = destino
                    escribir_log("SISTEMA", "COPIA_LOCAL", f"{nombre}: {exc}")
                    print(f"\nNo se pudo actualizar {nombre}; se usa la copia local válida: {exc}")
                else:
                    escribir_log("SISTEMA", "ERROR_DESCARGA", f"{nombre}: {exc}")
                    print(f"\nNo se pudo descargar {nombre} y no hay copia local válida: {exc}")
    return descargados


def _guardar_descarga(respuesta: requests.Response, temporal: Path) -> None:
    """Escribe y cierra una respuesta HTTP, aun cuando falle la transferencia."""
    with respuesta:
        respuesta.raise_for_status()
        with temporal.open("wb") as archivo:
            for bloque in respuesta.iter_content(chunk_size=64 * 1024):
                if bloque:
                    archivo.write(bloque)


def parse_fecha_manual(valor: object) -> pd.Timestamp | pd.NaT:
    if pd.isna(valor):
        return pd.NaT
    if isinstance(valor, (pd.Timestamp, datetime, np.datetime64)):
        fecha = pd.Timestamp(valor)
        return fecha if 1800 <= fecha.year <= datetime.now().year + 2 else pd.NaT

    texto = str(valor).strip()
    if not texto:
        return pd.NaT
    if re.fullmatch(r"\d{4}(?:\.0)?", texto):
        anio = int(float(texto))
        return pd.Timestamp(anio, 12, 31) if 1800 <= anio <= datetime.now().year + 2 else pd.NaT

    trimestre = re.fullmatch(r"(I|II|III|IV)[\s.-]+(\d{2}|\d{4})", texto, re.I)
    if trimestre:
        numero = {"I": 1, "II": 2, "III": 3, "IV": 4}[trimestre.group(1).upper()]
        texto_anio = trimestre.group(2)
        anio = int(texto_anio)
        if len(texto_anio) == 2:
            anio += 2000 if anio < 50 else 1900
        if not 1800 <= anio <= datetime.now().year + 2:
            return pd.NaT
        return pd.Period(year=anio, quarter=numero, freq="Q").end_time.normalize()

    mes_convertido = False
    for espanol in sorted(MESES, key=len, reverse=True):
        ingles = MESES[espanol]
        if texto.lower().startswith(espanol):
            texto = re.sub(rf"^{re.escape(espanol)}", ingles, texto, flags=re.I)
            mes_convertido = True
            break
    if mes_convertido:
        for formato in ("%b-%y", "%b %y", "%b.%y", "%b-%Y", "%b %Y", "%b.%Y"):
            try:
                return pd.to_datetime(texto, format=formato)
            except ValueError:
                continue
    try:
        fecha = pd.to_datetime(texto, format="mixed", dayfirst=True, errors="raise")
        return fecha if 1800 <= fecha.year <= datetime.now().year + 2 else pd.NaT
    except (TypeError, ValueError):
        return pd.NaT


def parse_fechas(fechas: list[object] | pd.Series) -> pd.DatetimeIndex:
    serie = pd.Series(fechas, dtype="object")
    resultado = pd.to_datetime(serie, format="mixed", dayfirst=True, errors="coerce")
    faltantes = resultado.isna() & serie.notna()
    if faltantes.any():
        resultado.loc[faltantes] = serie.loc[faltantes].map(parse_fecha_manual)
    return pd.DatetimeIndex(resultado)


def frecuencia_de_pestana(nombre: str) -> str:
    coincidencia = re.search(r"(?:^|\s)([ASTMD])$", nombre.strip(), re.I)
    return coincidencia.group(1).upper() if coincidencia else ""


def nombre_pestana(nombre: object) -> str:
    limpio = str(nombre).strip() or "Otros"
    return RENOMBRES_PESTANAS.get(limpio, limpio)[:31].strip()


def normalizar_fechas(fechas: pd.Series, frecuencia: str) -> pd.Series:
    fechas = pd.to_datetime(fechas, errors="coerce")
    if frecuencia == "A":
        return fechas.dt.to_period("Y").dt.start_time
    if frecuencia == "S":
        return fechas.map(
            lambda x: pd.Timestamp(x.year, 1 if x.month <= 6 else 7, 1)
            if pd.notna(x) else pd.NaT
        )
    if frecuencia == "T":
        return fechas.dt.to_period("Q").dt.start_time
    if frecuencia == "M":
        return fechas.dt.to_period("M").dt.start_time
    return fechas.dt.normalize()


def convertir_valor(valor: object) -> object:
    if pd.isna(valor):
        return np.nan
    if isinstance(valor, (int, float, np.number)) and not isinstance(valor, bool):
        return float(valor)
    texto = str(valor).strip()
    if not texto:
        return np.nan
    porcentaje = texto.endswith("%")
    if porcentaje:
        texto = texto[:-1].strip()
    if re.fullmatch(r"-?[\d.,]+", texto):
        if "," in texto:
            texto = texto.replace(".", "").replace(",", ".")
        try:
            return float(texto)
        except ValueError:
            pass
    return str(valor).strip()


def cargar_excel_completo(ruta: Path) -> dict[str, pd.DataFrame]:
    try:
        with pd.ExcelFile(ruta) as libro:
            return {
                hoja: pd.read_excel(libro, sheet_name=hoja, header=None)
                for hoja in libro.sheet_names
            }
    except Exception as exc:
        raise ValueError(f"No se pudo leer {ruta.name}: {exc}") from exc


def crear_indice_excel(excel_data: dict[str, pd.DataFrame]) -> dict[str, tuple[str, int, int]]:
    indice: dict[str, tuple[str, int, int]] = {}
    for hoja, dataframe in excel_data.items():
        for fila, valores in enumerate(dataframe.itertuples(index=False, name=None)):
            for columna, celda in enumerate(valores):
                if pd.isna(celda):
                    continue
                texto = str(celda).strip()
                if texto and texto not in indice:
                    indice[texto] = (hoja, fila, columna)
    return indice


def _columna_fecha(dataframe: pd.DataFrame, fila_id: int, columna_id: int) -> int:
    candidatas = range(min(3, dataframe.shape[1], max(columna_id, 1)))
    mejor_columna, mejor_puntaje = 0, -1
    muestra = dataframe.iloc[fila_id + 1 : fila_id + 31]
    for columna in candidatas:
        puntaje = sum(pd.notna(parse_fecha_manual(valor)) for valor in muestra.iloc[:, columna])
        if puntaje > mejor_puntaje:
            mejor_columna, mejor_puntaje = columna, puntaje
    return mejor_columna


def _tipo_fecha_cruda(valor: object) -> str:
    """Distingue bloques anuales/trimestrales pegados dentro de una hoja."""
    if isinstance(valor, (pd.Timestamp, datetime, np.datetime64)):
        return "fecha"
    texto = str(valor).strip()
    if re.fullmatch(r"\d{4}(?:\.0)?", texto):
        return "anio"
    if re.fullmatch(r"(I|II|III|IV)[\s.-]+(\d{2}|\d{4})", texto, re.I):
        return "trimestre"
    if any(texto.lower().startswith(mes) for mes in MESES):
        return "mes"
    return "fecha"


def extraer_serie_desde_indice(
    excel_data: dict[str, pd.DataFrame], ubicacion: tuple[str, int, int]
) -> pd.DataFrame:
    """Extrae una tabla y se detiene al terminar el bloque que contiene el ID."""
    hoja, fila_id, columna_id = ubicacion
    dataframe = excel_data[hoja]
    columna_fecha = _columna_fecha(dataframe, fila_id, columna_id)

    fila_inicio = None
    for fila in range(fila_id + 1, min(fila_id + 1001, len(dataframe))):
        if pd.notna(parse_fecha_manual(dataframe.iat[fila, columna_fecha])):
            fila_inicio = fila
            break
    if fila_inicio is None:
        raise ValueError(f"No hay fechas debajo del ID en la hoja {hoja}")

    tipo_bloque = _tipo_fecha_cruda(dataframe.iat[fila_inicio, columna_fecha])
    fechas: list[object] = []
    valores: list[object] = []
    for fila in range(fila_inicio, len(dataframe)):
        fecha_cruda = dataframe.iat[fila, columna_fecha]
        if tipo_bloque in {"anio", "trimestre", "mes"} and _tipo_fecha_cruda(fecha_cruda) != tipo_bloque:
            break
        fecha = parse_fecha_manual(fecha_cruda)
        if pd.isna(fecha):
            break
        fechas.append(fecha)
        valores.append(convertir_valor(dataframe.iat[fila, columna_id]))

    serie = pd.DataFrame({"fecha": pd.DatetimeIndex(fechas), "valor": valores})
    serie = serie[serie["fecha"] <= pd.Timestamp.today().normalize()]
    con_valor = serie["valor"].notna()
    if not con_valor.any():
        raise ValueError(f"El ID encontrado en {hoja} no contiene valores")
    primero, ultimo = con_valor[con_valor].index[[0, -1]]
    serie = serie.loc[primero:ultimo].copy()
    serie = compactar_fechas(serie)
    if serie.empty:
        raise ValueError("No se extrajeron datos válidos")
    return serie


def _ultimo_no_nulo(serie: pd.Series) -> object:
    valores = serie.dropna()
    return valores.iloc[-1] if not valores.empty else np.nan


def compactar_fechas(dataframe: pd.DataFrame) -> pd.DataFrame:
    """Une fechas repetidas conservando el valor no vacío más reciente."""
    if dataframe.empty or "fecha" not in dataframe:
        return dataframe
    dataframe = dataframe.dropna(subset=["fecha"]).copy()
    columnas = [columna for columna in dataframe.columns if columna != "fecha"]
    if dataframe["fecha"].duplicated().any():
        dataframe = dataframe.groupby("fecha", as_index=False, sort=True)[columnas].agg(_ultimo_no_nulo)
    return dataframe.sort_values("fecha").reset_index(drop=True)


def preparar_hoja_bd(dataframe: pd.DataFrame, pestana: str) -> pd.DataFrame:
    dataframe = dataframe.copy()
    if "fecha" not in dataframe:
        dataframe.insert(0, "fecha", pd.NaT)
    dataframe["fecha"] = normalizar_fechas(dataframe["fecha"], frecuencia_de_pestana(pestana))
    hoy = pd.Timestamp.today().normalize()
    dataframe = dataframe[
        dataframe["fecha"].notna()
        & (dataframe["fecha"].dt.year >= 1800)
        & (dataframe["fecha"] <= hoy)
    ]
    dataframe = compactar_fechas(dataframe)
    columnas_datos = [columna for columna in dataframe if columna != "fecha"]
    if columnas_datos:
        dataframe = dataframe.dropna(how="all", subset=columnas_datos)
    return dataframe


def actualizar_serie(
    existente: pd.DataFrame, nueva: pd.DataFrame, variable: str, pestana: str
) -> pd.DataFrame:
    existente = preparar_hoja_bd(existente, pestana)
    nueva = nueva.rename(columns={"valor": variable})
    nueva["fecha"] = normalizar_fechas(nueva["fecha"], frecuencia_de_pestana(pestana))
    nueva = compactar_fechas(nueva)

    if variable not in existente:
        resultado = pd.merge(existente, nueva, on="fecha", how="outer")
    else:
        fecha_corte = nueva["fecha"].min()
        anterior = existente[existente["fecha"] < fecha_corte]
        actual = existente[existente["fecha"] >= fecha_corte].drop(columns=[variable])
        resultado = pd.concat(
            [anterior, pd.merge(actual, nueva, on="fecha", how="outer")], ignore_index=True
        )
    return preparar_hoja_bd(resultado, pestana)


def aplicar_formatos_fecha(ruta: Path) -> None:
    """Muestra fechas sin hora conservando celdas de fecha comparables."""
    libro = load_workbook(ruta)
    formatos = {"A": "yyyy", "S": "yyyy-mm", "T": "yyyy-mm", "M": "yyyy-mm", "D": "yyyy-mm-dd"}
    for hoja in libro.worksheets:
        if hoja.title in HOJAS_ADMINISTRATIVAS or hoja.title in HOJAS_AJENAS_IED:
            continue
        formato = formatos.get(frecuencia_de_pestana(hoja.title), "yyyy-mm-dd")
        for celda in hoja["A"][1:]:
            if celda.value is not None:
                celda.number_format = formato
        hoja.column_dimensions["A"].width = max(10, len("fecha") + 2)
    libro.save(ruta)


def _validar_guardado(ruta: Path, hojas: dict[str, pd.DataFrame], formatos: dict[str, str]) -> None:
    """Reabre el resultado y comprueba dimensiones, fechas y filas vacías."""
    libro = load_workbook(ruta, read_only=True, data_only=True)
    try:
        for nombre, datos in hojas.items():
            hoja = libro[nombre]
            if hoja.max_row != len(datos) + 1 or hoja.max_column != len(datos.columns):
                raise ValueError(
                    f"Dimensiones incorrectas en {nombre}: "
                    f"{hoja.max_row}x{hoja.max_column}, esperado {len(datos) + 1}x{len(datos.columns)}"
                )
            if str(hoja.cell(1, 1).value).strip().casefold() != "fecha":
                raise ValueError(f"La primera columna de {nombre} no es fecha")
            for fila in hoja.iter_rows(min_row=2, values_only=False):
                if all(celda.value is None for celda in fila):
                    raise ValueError(f"Fila completamente vacía en {nombre}, fila {fila[0].row}")
                if fila[0].value is None:
                    raise ValueError(f"Fecha vacía en {nombre}, fila {fila[0].row}")
                if fila[0].number_format != formatos[nombre]:
                    raise ValueError(f"Formato de fecha incorrecto en {nombre}, fila {fila[0].row}")
    finally:
        libro.close()


def guardar_datos_preservando_formato(
    ruta: Path,
    hojas: dict[str, pd.DataFrame],
    frecuencias: dict[str, str] | None = None,
    hojas_obsoletas: set[str] | None = None,
) -> None:
    """Recrea hojas limpias y publica el libro sólo después de validarlo."""
    libro = load_workbook(ruta)
    temporal = ruta.with_name(f"{ruta.stem}.actualizacion{ruta.suffix}")
    formatos = {"A": "yyyy", "S": "yyyy-mm", "T": "yyyy-mm", "M": "yyyy-mm", "D": "yyyy-mm-dd"}
    frecuencias = frecuencias or {nombre: frecuencia_de_pestana(nombre) for nombre in hojas}
    formatos_hojas = {nombre: formatos.get(frecuencias.get(nombre, ""), "yyyy-mm-dd") for nombre in hojas}

    def valor_excel(valor: object) -> object:
        if pd.isna(valor):
            return None
        if isinstance(valor, pd.Timestamp):
            return valor.to_pydatetime()
        if isinstance(valor, np.generic):
            return valor.item()
        return valor

    for nombre in sorted((hojas_obsoletas or set()) - set(hojas)):
        if nombre in libro.sheetnames and nombre not in HOJAS_ADMINISTRATIVAS | HOJAS_AJENAS_IED:
            del libro[nombre]

    for nombre, datos in hojas.items():
        nombre_hoja = str(nombre).strip()[:31]
        if nombre_hoja in libro.sheetnames:
            del libro[nombre_hoja]
        hoja = libro.create_sheet(nombre_hoja)
        columnas = list(datos.columns)
        filas_nuevas = len(datos) + 1

        for columna, encabezado in enumerate(columnas, 1):
            hoja.cell(1, columna).value = encabezado
        for indice, fila in enumerate(datos.itertuples(index=False, name=None), 2):
            for columna, valor in enumerate(fila, 1):
                hoja.cell(indice, columna).value = valor_excel(valor)
        formato = formatos_hojas[nombre]
        for celda in hoja["A"][1:filas_nuevas]:
            if celda.value is not None:
                celda.number_format = formato
        hoja.freeze_panes = "B2"

    try:
        libro.save(temporal)
        libro.close()
        _validar_guardado(temporal, hojas, formatos_hojas)
        temporal.replace(ruta)
    except Exception:
        libro.close()
        temporal.unlink(missing_ok=True)
        raise


def procesar_datos(descargar: bool = True) -> dict[str, int]:
    """Ejecuta el flujo completo del IED y actualiza BD.xlsx."""
    escribir_log("SISTEMA", "INICIO", "Inicio de la extracción IED")
    if descargar:
        archivos = descargar_excels()
    else:
        archivos = {nombre: CARPETA_FUENTE / nombre for nombre in EXCEL_URLS if (CARPETA_FUENTE / nombre).exists()}
    faltantes = set(EXCEL_URLS) - set(archivos)
    if faltantes:
        raise RuntimeError(f"Faltan libros del IED: {', '.join(sorted(faltantes))}")

    libros: dict[str, dict[str, pd.DataFrame]] = {}
    indices: dict[str, dict[str, tuple[str, int, int]]] = {}
    for nombre, ruta in tqdm(archivos.items(), desc="Cargando libros IED"):
        libros[nombre] = cargar_excel_completo(ruta)
        indices[nombre] = crear_indice_excel(libros[nombre])

    existentes = load_existing_catalog(ARCHIVO_BD)
    hojas_obsoletas = set(existentes.get("Pestaña BD", pd.Series(dtype=str)).dropna().astype(str))
    hojas_salida, inventario, errores = discover_and_extract(
        libros, indices, existentes, extraer_serie_desde_indice, normalizar_fechas
    )
    frecuencias = {
        pestana: str(grupo["Frecuencia"].iloc[0])
        for pestana, grupo in inventario.groupby("Pestaña BD")
    }
    guardar_datos_preservando_formato(
        ARCHIVO_BD, hojas_salida, frecuencias=frecuencias, hojas_obsoletas=hojas_obsoletas
    )
    from tools.generar_codificacion import generar as generar_codificacion
    generar_codificacion(inventario)

    for id_serie, error in errores:
        escribir_log(id_serie, "ERROR", error)
    resumen = {"total": len(inventario) + len(errores), "exitosas": len(inventario), "fallidas": len(errores)}
    escribir_log("SISTEMA", "FIN", str(resumen))
    print(f"\nIED terminado: {len(inventario)}/{resumen['total']} series; errores: {len(errores)}")
    return resumen


def ejecutar() -> None:
    resumen = procesar_datos(descargar=True)
    if resumen["fallidas"]:
        raise RuntimeError(
            f"IED terminó con {resumen['fallidas']} de {resumen['total']} series fallidas; "
            "consulte el log para ver el detalle"
        )
