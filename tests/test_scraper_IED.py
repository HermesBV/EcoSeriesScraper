import unittest
from pathlib import Path
from unittest.mock import patch

import pandas as pd
from openpyxl import Workbook, load_workbook

from scrapers import scraper_IED


class RespuestaFalsa:
    def __init__(self, bloques: list[bytes]) -> None:
        self.bloques = bloques
        self.cerrada = False

    def __enter__(self):
        return self

    def __exit__(self, *args):
        self.cerrada = True

    def raise_for_status(self) -> None:
        return None

    def iter_content(self, chunk_size: int):
        return iter(self.bloques)


class ScraperIEDTests(unittest.TestCase):
    def test_guardar_descarga_escribe_bloques_y_cierra_respuesta(self) -> None:
        destino = Path(__file__).parent / "_descarga_prueba.xlsx"
        self.addCleanup(destino.unlink, missing_ok=True)
        respuesta = RespuestaFalsa([b"abc", b"", b"def"])

        scraper_IED._guardar_descarga(respuesta, destino)

        self.assertEqual(destino.read_bytes(), b"abcdef")
        self.assertTrue(respuesta.cerrada)

    def test_guardado_preserva_hojas_administrativas_y_reemplaza_atomico(self) -> None:
        ruta = Path(__file__).parent / "_BD_prueba.xlsx"
        temporal = Path(__file__).parent / "_BD_prueba.actualizacion.xlsx"
        self.addCleanup(ruta.unlink, missing_ok=True)
        self.addCleanup(temporal.unlink, missing_ok=True)
        libro = Workbook()
        libro.active.title = "Codificacion"
        libro["Codificacion"]["A1"] = "control"
        libro.create_sheet("Serie M")
        libro.save(ruta)

        datos = pd.DataFrame(
            {"fecha": [pd.Timestamp(2026, 8, 1)], "variable": [12.5]}
        )
        scraper_IED.guardar_datos_preservando_formato(ruta, {"Serie M": datos})

        actualizado = load_workbook(ruta)
        self.assertEqual(actualizado["Codificacion"]["A1"].value, "control")
        self.assertEqual(actualizado["Serie M"]["B2"].value, 12.5)
        self.assertEqual(actualizado["Serie M"]["A2"].number_format, "yyyy-mm")
        actualizado.close()
        self.assertFalse(temporal.exists())

    def test_ejecutar_informa_series_fallidas(self) -> None:
        resumen = {"total": 3, "exitosas": 2, "fallidas": 1}
        with patch.object(scraper_IED, "procesar_datos", return_value=resumen):
            with self.assertRaisesRegex(RuntimeError, "1 de 3 series fallidas"):
                scraper_IED.ejecutar()


if __name__ == "__main__":
    unittest.main()
