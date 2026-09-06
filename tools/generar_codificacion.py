"""Actualiza el inventario maestro de series dentro de BD.xlsx."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
DB_FILE = ROOT / "BD.xlsx"
TEMP_FILE = ROOT / ".BD_codificacion.tmp.xlsx"
CODE_SHEETS = {
    "Referencia_Codigos", "Parentesco_Codigos", "Introduccion_Codigos",
    "Mapa_Tematico", "Referencias",
}
INVENTORY_COLUMNS = [
    "ID", "Código fuente", "ID origen", "Nombre serie", "Variable", "Unidades", "Valoración", "Descripción",
    "Frecuencia", "Pestaña BD", "Columna BD", "Archivo origen", "Hoja origen",
    "Origen", "Fuente", "Catálogo ID", "Dataset ID", "Distribución ID",
    "Título dataset", "Tema dataset", "Responsable dataset", "Fuente de valores",
    "Fecha inicio", "Fecha fin", "Estado",
]


def _load_current_inventory() -> pd.DataFrame:
    with pd.ExcelFile(DB_FILE) as book:
        if "Codificacion" not in book.sheet_names:
            return pd.DataFrame(columns=INVENTORY_COLUMNS)
        data = pd.read_excel(book, sheet_name="Codificacion", dtype=object)
    return data if "ID" in data.columns else pd.DataFrame(columns=INVENTORY_COLUMNS)


def _communication_row() -> dict[str, object]:
    return {
        "ID": "bcra::comunicaciones-A-B-C-P-desde-2006",
        "Código fuente": "bcra",
        "ID origen": "comunicaciones-A-B-C-P-desde-2006",
        "Nombre serie": "Comunicaciones BCRA",
        "Variable": "Comunicaciones BCRA tipos A, B, C y P",
        "Unidades": "Documentos",
        "Valoración": "No aplica / no informado",
        "Descripción": "Inventario de comunicaciones publicadas por el BCRA.",
        "Frecuencia": "I",
        "Pestaña BD": "Comunicaciones BCRA",
        "Columna BD": "",
        "Archivo origen": "",
        "Hoja origen": "",
        "Origen": "BCRA: Buscador de comunicaciones",
        "Fuente": "https://www.bcra.gob.ar/buscador-de-comunicaciones/",
        "Catálogo ID": "bcra",
        "Fuente de valores": "API BCRA",
        "Fecha inicio": pd.Timestamp(2006, 1, 1),
        "Fecha fin": None,
        "Estado": "VIGENTE",
    }


def _normalize_inventory(inventory: pd.DataFrame) -> pd.DataFrame:
    result = inventory.copy()
    if "ID origen" not in result:
        raise ValueError("El inventario no contiene 'ID origen'")
    result["ID origen"] = result["ID origen"].astype(str).str.strip()
    if "Código fuente" not in result:
        result["Código fuente"] = "datos.gob.ar"
    result["Código fuente"] = result["Código fuente"].astype(str).str.strip()
    result["ID"] = result["Código fuente"] + "::" + result["ID origen"]
    result = result[result["ID origen"].ne("") & result["ID origen"].ne("nan")]
    result = result.drop_duplicates(["Código fuente", "ID origen"], keep="last")
    for column in INVENTORY_COLUMNS:
        if column not in result:
            result[column] = None
    return result[INVENTORY_COLUMNS].sort_values(
        ["Archivo origen", "Hoja origen", "ID"], na_position="last"
    )


def _write_inventory(sheet, inventory: pd.DataFrame) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for column, name in enumerate(INVENTORY_COLUMNS, 1):
        cell = sheet.cell(1, column, name)
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for row_number, row in enumerate(inventory.itertuples(index=False, name=None), 2):
        for column, value in enumerate(row, 1):
            sheet.cell(row_number, column, None if pd.isna(value) else value)
        for column in (23, 24):
            if sheet.cell(row_number, column).value is not None:
                sheet.cell(row_number, column).number_format = "yyyy-mm-dd"
    if len(inventory):
        table = Table(
            displayName="TablaInventarioSeries",
            ref=f"A1:{get_column_letter(len(INVENTORY_COLUMNS))}{len(inventory) + 1}",
        )
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        sheet.add_table(table)
    sheet.freeze_panes = "A2"
    for column in range(1, len(INVENTORY_COLUMNS) + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 22

def generar(inventory: pd.DataFrame | None = None) -> None:
    data = _load_current_inventory() if inventory is None else inventory
    if "Pestaña BD" in data:
        data = data[data["Pestaña BD"].ne("Comunicaciones BCRA")]
    book = load_workbook(DB_FILE)
    if "Comunicaciones BCRA" in book.sheetnames:
        data = pd.concat([data, pd.DataFrame([_communication_row()])], ignore_index=True)
    data = _normalize_inventory(data)

    for name in CODE_SHEETS:
        if name in book.sheetnames:
            del book[name]
    if "Codificacion" in book.sheetnames:
        del book["Codificacion"]
    sheet = book.create_sheet("Codificacion", 0)
    _write_inventory(sheet, data)

    try:
        book.save(TEMP_FILE)
        book.close()
        check = pd.read_excel(TEMP_FILE, sheet_name="Codificacion")
        if len(check) != len(data) or check["ID"].duplicated().any():
            raise ValueError("La validación del inventario guardado falló")
        TEMP_FILE.replace(DB_FILE)
    finally:
        TEMP_FILE.unlink(missing_ok=True)


if __name__ == "__main__":
    generar()
