import pandas as pd
import os
import numpy as np
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import requests
import urllib3
from tqdm import tqdm
import warnings

# Configuración inicial
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
warnings.simplefilter(action='ignore', category=FutureWarning)

# Diccionario de URLs y nombres de archivos
EXCEL_URLS = {
    "empleo_ingresos.xlsx": "https://www.economia.gob.ar/download/infoeco/apendice3a.xlsx",
    "sector_externo.xlsx": "https://www.economia.gob.ar/download/infoeco/apendice5.xlsx",
    "internacional.xlsx": "https://www.economia.gob.ar/download/infoeco/internacional_ied.xlsx",
    "dinero_bancos.xlsx": "https://www.economia.gob.ar/download/infoeco/apendice8.xlsx",
    "precios.xlsx": "https://www.economia.gob.ar/download/infoeco/apendice4.xlsx",
    "finanzas.xlsx": "https://www.economia.gob.ar/download/infoeco/apendice-financiero.xlsx",
    "finanzas_publicas.xlsx": "https://www.economia.gob.ar/download/infoeco/apendice6.xlsx",
    "actividad.xlsx": "https://www.economia.gob.ar/download/infoeco/actividad_ied.xlsx"
}

# --- LIBRERÍA DE MAPEO ---
# Conecta el valor de la columna 'Excel Origen' (en Codigos.xlsx) 
# con el nombre del archivo que descarga el script (en EXCEL_URLS).
# **Asegúrate de que las claves aquí coincidan con tu Excel 'Codigos.xlsx'**

MAPEO_ORIGEN_ARCHIVO = {
    # Mapeos existentes que funcionaban
    "Empleo e Ingresos: Apendice3A": "empleo_ingresos.xlsx",
    "Actividad: Actividad_IED": "actividad.xlsx",
    "Contexto Internacional": "internacional.xlsx",
    "Precios: Apendice4": "precios.xlsx",
    "Finanzas: Apendice-Financiero": "finanzas.xlsx",
    "Finanzas Públicas: Apendice6": "finanzas_publicas.xlsx",
    "Sector Externo: Apendice5": "sector_externo.xlsx",
    "Dinero y Bancos: Apendice8": "dinero_bancos.xlsx",
    
    # Mapeos alternativos por si acaso (redundancia)
    "Precios": "precios.xlsx",
    "Finanzas": "finanzas.xlsx",
    "Finanzas Públicas": "finanzas_publicas.xlsx",
    "Sector Externo": "sector_externo.xlsx",
    "Dinero y Bancos": "dinero_bancos.xlsx"
}

def limpiar_nombres_definidos(ruta_archivo):
    """
    Elimina los 'Defined Names' (rangos con nombre) que causan conflictos
    al abrir los archivos de gobierno en Excel.
    """
    try:
        # Cargar sin leer datos, solo estructura, para ser rápido
        wb = load_workbook(ruta_archivo)
        
        # Lista de nombres a eliminar
        nombres = list(wb.defined_names.keys())
        
        if nombres:
            for nombre in nombres:
                # Eliminar el nombre definido
                del wb.defined_names[nombre]
            
            # Guardar el archivo limpio
            wb.save(ruta_archivo)
            # print(f"Limpiados {len(nombres)} nombres conflictivos en {os.path.basename(ruta_archivo)}")
            
    except Exception as e:
        # Si falla (ej. formato muy viejo), lo ignoramos y seguimos
        pass

# --- DÓNDE PONERLO EN TU CÓDIGO ---
# Dentro de la función descargar_excels(), justo después del 'with open...'
def crear_carpeta_logs():
    """Crea la carpeta de logs si no existe"""
    if not os.path.exists('logs'):
        os.makedirs('logs')

def crear_carpeta_excels():
    """Crea la carpeta para los excels si no existe"""
    if not os.path.exists('Excels_IED'):
        os.makedirs('Excels_IED')

def limpiar_nombres_definidos(ruta_archivo):
    """
    Elimina los 'Defined Names' (rangos con nombre) que causan conflictos
    al abrir los archivos de gobierno en Excel.
    """
    try:
        # Cargar sin leer datos, solo estructura, para ser rápido
        wb = load_workbook(ruta_archivo)
        
        # Lista de nombres a eliminar
        nombres = list(wb.defined_names.keys())
        
        if nombres:
            for nombre in nombres:
                # Eliminar el nombre definido
                del wb.defined_names[nombre]
            
            # Guardar el archivo limpio
            wb.save(ruta_archivo)
            # print(f"Limpiados {len(nombres)} nombres conflictivos en {os.path.basename(ruta_archivo)}")
            
    except Exception as e:
        # Si falla (ej. formato muy viejo), lo ignoramos y seguimos
        pass

# --- DÓNDE PONERLO EN TU CÓDIGO ---
# Dentro de la función descargar_excels(), justo después del 'with open...'

def descargar_excels():
    """Descarga todos los archivos Excel sobrescribiendo los existentes"""
    crear_carpeta_excels()
    descargados = {}
    
    for nombre_archivo, url in tqdm(EXCEL_URLS.items(), desc="Descargando archivos"):
        destino = os.path.join('Excels_IED', nombre_archivo)
        
        try:
            response = requests.get(url, stream=True, verify=False, timeout=30)
            response.raise_for_status()
            
            with open(destino, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
            
            # --- NUEVO: LIMPIEZA ---
            # Limpiamos el archivo justo después de bajarlo
            limpiar_nombres_definidos(destino)
            # -----------------------

            descargados[nombre_archivo] = destino
            
        except Exception as e:
            error_msg = f"Error al descargar {nombre_archivo}: {str(e)}"
            print(f"\n{error_msg}")
            escribir_log("SISTEMA", "ERROR_DESCARGA", error_msg)
    
    return descargados

def escribir_log(id_serie, estado, mensaje=""):
    """Escribe un mensaje en el archivo de log"""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    log_file = f"logs/log_{datetime.now().strftime('%Y%m%d')}.txt"
    with open(log_file, 'a', encoding='utf-8') as f:
        f.write(f"{timestamp}|{id_serie}|{estado}")
        if mensaje:
            mensaje_limpio = mensaje.replace('\n', ' || ')
            f.write(f"|{mensaje_limpio}")
        f.write("\n")

def parse_fechas(fechas):
    """Función mejorada para el parseo de fechas con manejo específico de formatos"""
    # Primero intentamos con formatos conocidos
    formatos_conocidos = [
        '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y',
        '%b-%y', '%b %Y', '%m-%Y',
        '%Y',
    ]
    
    # Intento de conversión masiva primero (mucho más rápido)
    try:
        fechas_dt = pd.to_datetime(fechas, errors='coerce')
        # Si la mayoría no son NaT, usamos este resultado
        if fechas_dt.notna().mean() > 0.8:
             # Forzar fin de mes para los que parecen ser mensuales/trimestrales/anuales
            for i, fecha in enumerate(fechas_dt):
                if not pd.isna(fecha) and fecha.day == 1:
                    fechas_dt[i] = fecha + pd.offsets.MonthEnd(0)
            return fechas_dt
    except Exception:
        pass # Continuar con el parseo manual si falla

    # Si no funciona con formatos conocidos o falló la conversión masiva, aplicamos el parser manual
    fechas_parseadas = []
    for fecha in fechas:
        fecha_parseada = parse_fecha_manual(fecha)
        fechas_parseadas.append(fecha_parseada if not pd.isna(fecha_parseada) else pd.NaT)
    
    return pd.DatetimeIndex(fechas_parseadas)

def parse_fecha_manual(s):
    """Parser manual para formatos de fecha no estándar"""
    try:
        s = str(s).strip()
        
        # Caso 0: Ya es una fecha (ej. 1993-01-31 00:00:00)
        if re.match(r'^\d{4}-\d{2}-\d{2}', s):
            return pd.to_datetime(s)
            
        # Caso 1: año de 4 dígitos
        if re.match(r'^\d{4}$', s):
            year = int(s)
            return pd.Timestamp(year=year, month=12, day=31)
            
        # --- NUEVO CASO AGREGADO: Trimestre con guion (I-92, IV-24) ---
        match_trim_guion = re.match(r'^(I|II|III|IV)-(\d{2}|\d{4})$', s, re.IGNORECASE)
        if match_trim_guion:
            trimestre = match_trim_guion.group(1).upper()
            año_str = match_trim_guion.group(2)
            año = int(año_str)
            if len(año_str) == 2:
                año = 2000 + año if año < 50 else 1900 + año 
                
            if trimestre == 'I': return pd.Timestamp(año, 3, 31)
            elif trimestre == 'II': return pd.Timestamp(año, 6, 30)
            elif trimestre == 'III': return pd.Timestamp(año, 9, 30)
            elif trimestre == 'IV': return pd.Timestamp(año, 12, 31)
        # -------------------------------------------------------------

        # Caso 2: trimestre con espacio (I 24, II 24)
        match_trim = re.match(r'^(I|II|III|IV)\s+(\d{2}|\d{4})$', s, re.IGNORECASE)
        if match_trim:
            trimestre = match_trim.group(1).upper()
            año_str = match_trim.group(2)
            año = int(año_str)
            if len(año_str) == 2:
                año = 2000 + año if año < 50 else 1900 + año
                
            if trimestre == 'I': return pd.Timestamp(año, 3, 31)
            elif trimestre == 'II': return pd.Timestamp(año, 6, 30)
            elif trimestre == 'III': return pd.Timestamp(año, 9, 30)
            elif trimestre == 'IV': return pd.Timestamp(año, 12, 31)

        # Caso 3: Trimestre con punto (IV.02, I.23)
        match_trim_punto = re.match(r'^(I|II|III|IV)\.(\d{2}|\d{4})$', s, re.IGNORECASE)
        if match_trim_punto:
            trimestre = match_trim_punto.group(1).upper()
            año_corto = int(match_trim_punto.group(2))
            # Manejo si viene 2 o 4 digitos en el punto
            if len(str(año_corto)) == 4:
                año = año_corto
            else:
                año = 2000 + año_corto if año_corto < 50 else 1900 + año_corto
            
            if trimestre == 'I': return pd.Timestamp(año, 3, 31)
            elif trimestre == 'II': return pd.Timestamp(año, 6, 30)
            elif trimestre == 'III': return pd.Timestamp(año, 9, 30)
            elif trimestre == 'IV': return pd.Timestamp(año, 12, 31)

        # Caso 4: meses (Ene-24, Ene 24, etc.)
        meses_esp = {
            'Ene': 'Jan', 'Feb': 'Feb', 'Mar': 'Mar', 'Abr': 'Apr', 
            'May': 'May', 'Jun': 'Jun', 'Jul': 'Jul', 'Ago': 'Aug', 
            'Sep': 'Sep', 'Oct': 'Oct', 'Nov': 'Nov', 'Dic': 'Dec',
            'Enero': 'Jan', 'Febrero': 'Feb', 'Marzo': 'Mar', 'Abril': 'Apr',
            'Mayo': 'May', 'Junio': 'Jun', 'Julio': 'Jul', 'Agosto': 'Aug',
            'Septiembre': 'Sep', 'Octubre': 'Oct', 'Noviembre': 'Nov', 'Diciembre': 'Dec'
        }
        for mes_esp, mes_eng in meses_esp.items():
            if s.lower().startswith(mes_esp.lower()):
                s_eng = re.sub(f'^{mes_esp}', mes_eng, s, flags=re.IGNORECASE)
                # Intentar varios formatos, agregando soporte para guiones extraños
                for fmt in ['%b %y', '%b-%y', '%b %Y', '%b-%Y', '%b.%y', '%b.%Y']:
                    try:
                        fecha = pd.to_datetime(s_eng, format=fmt)
                        return fecha + pd.offsets.MonthEnd(0)
                    except ValueError:
                        continue
                        
        # Caso 5: Formatos estándar (último recurso)
        return pd.to_datetime(s, dayfirst=True, errors='coerce')
        
    except Exception:
        return pd.NaT

def crear_indice_excel(excel_data):
    """
    Crea un índice (diccionario) de todos los valores en todas las hojas 
    de un archivo Excel para búsqueda rápida.
    El índice mapea: valor -> (nombre_hoja, fila, columna)
    """
    indice = {}
    for sheet_name, df in excel_data.items():
        # Iterar sobre el DataFrame (más rápido que itertuples/iterrows para numpy)
        for r_idx, fila in enumerate(df.values):
            for c_idx, celda in enumerate(fila):
                valor = str(celda).strip()
                if valor and valor not in indice:
                    # Guardar la ubicación 0-based del DataFrame
                    indice[valor] = (sheet_name, r_idx, c_idx)
    return indice

def extraer_serie_desde_indice(excel_data, ubicacion):
    """
    Extrae una serie de datos (fechas, valores) usando la ubicación exacta.
    Detecta automáticamente si la fecha está en Columna A (0) o B (1).
    """
    sheet_name, fila_id, col_id = ubicacion
    df = excel_data[sheet_name]
    
    fechas = []
    valores = []
    
    # --- 1. DETECCIÓN INTELIGENTE DE COLUMNA DE FECHA ---
    # A veces la Columna A (0) está vacía y la fecha está en la B (1).
    # Revisamos las 20 filas siguientes al ID para ver cuál tiene datos.
    col_fecha = 0 # Por defecto columna A
    
    conteo_col0 = 0
    conteo_col1 = 0
    start_check = fila_id + 1
    end_check = min(fila_id + 25, df.shape[0])
    
    for r_check in range(start_check, end_check):
        val0 = df.iat[r_check, 0]
        val1 = df.iat[r_check, 1]
        
        if pd.notna(val0) and str(val0).strip() != "":
            conteo_col0 += 1
        if pd.notna(val1) and str(val1).strip() != "":
            conteo_col1 += 1
            
    # Si la columna A está vacía y la B tiene datos, cambiamos a la B
    if conteo_col0 == 0 and conteo_col1 > 0:
        col_fecha = 1
        # print(f"   -> Detectado: Fechas en Columna B para {sheet_name}")

    # --- 2. ENCONTRAR EL INICIO DE LOS DATOS ---
    fila_inicio_datos = -1
    # Aumentamos el rango de búsqueda por si hay mucho espacio vacío
    max_filas_a_buscar = 1000 
    
    for r_offset in range(1, max_filas_a_buscar + 1):
        r = fila_id + r_offset
        if r >= df.shape[0]:
            break 

        # Miramos la columna de fecha detectada
        fecha_val = df.iat[r, col_fecha]
        
        # Si hay algo en la celda de fecha, asumimos que aquí arrancan los datos
        if (pd.notna(fecha_val) and str(fecha_val).strip() != ""):
            fila_inicio_datos = r
            break
            
    if fila_inicio_datos == -1:
        raise ValueError(f"No se encontraron datos válidos debajo del ID en {sheet_name} (Buscando fecha en Columna {col_fecha})")

    # --- 3. EXTRACCIÓN Y LIMPIEZA (CON SOPORTE DE PORCENTAJES) ---
    r = fila_inicio_datos
    while r < df.shape[0]:
        fecha_val = df.iat[r, col_fecha]
        
        # Condición de parada: fecha vacía
        if pd.isna(fecha_val) or str(fecha_val).strip() == "":
            # Pequeño chequeo: si la fila siguiente SÍ tiene fecha, no paramos (es un hueco)
            if r + 1 < df.shape[0]:
                next_val = df.iat[r+1, col_fecha]
                if pd.notna(next_val) and str(next_val).strip() != "":
                    r += 1
                    continue # Saltamos esta fila vacía
            break # Si no, terminamos

        valor_val = df.iat[r, col_id] 
        
        fechas.append(fecha_val)
        
        # Limpieza de Valor (Porcentajes, Miles, etc.)
        if pd.isna(valor_val):
            valores.append(np.nan)
        
        elif isinstance(valor_val, (int, float)):
            valores.append(float(valor_val))
            
        else:
            valor_val_limpio = str(valor_val).strip()
            
            # Caso Porcentaje (ej: "38,4%")
            if '%' in valor_val_limpio:
                try:
                    num_str = valor_val_limpio.replace('%', '').strip()
                    # Si tiene coma, asumimos formato europeo/latam
                    if ',' in num_str:
                        num_str = num_str.replace('.', '')  # Quitar punto de miles
                        num_str = num_str.replace(',', '.') # Coma a punto
                    valores.append(float(num_str))
                except:
                    valores.append(np.nan)
            
            # Caso Número Texto con coma (ej: "1.234,56")
            elif re.match(r'^-?[\d\.,]+$', valor_val_limpio):
                try:
                    # Si tiene punto y coma, el punto es miles y la coma decimal
                    if '.' in valor_val_limpio and ',' in valor_val_limpio:
                         num_str = valor_val_limpio.replace('.', '').replace(',', '.')
                    # Si solo tiene coma, es decimal
                    elif ',' in valor_val_limpio:
                         num_str = valor_val_limpio.replace(',', '.')
                    else:
                         num_str = valor_val_limpio
                    
                    valores.append(float(num_str))
                except:
                    valores.append(np.nan)

            elif valor_val_limpio == "":
                valores.append(np.nan)
            else:
                valores.append(valor_val_limpio) # Guardar texto tal cual (s/d, etc)
        
        r += 1
    
    if not fechas:
        raise ValueError("No se extrajo ninguna fecha/valor")
    
    fechas_dt = parse_fechas(fechas)
    
    df_temp = pd.DataFrame({'fecha': fechas_dt, 'valor': valores})
    df_temp = df_temp.dropna(subset=['fecha'])
    
    return df_temp['fecha'], df_temp['valor']


def cargar_excel_completo(ruta):
    """Carga un archivo Excel completo con todas sus hojas"""
    excel_data = {}
    try:
        with pd.ExcelFile(ruta) as xls:
            for sheet_name in xls.sheet_names:
                excel_data[sheet_name] = pd.read_excel(xls, sheet_name, header=None)
        return excel_data
    except Exception as e:
        raise ValueError(f"Error al cargar archivo {os.path.basename(ruta)}: {str(e)}")

def procesar_datos():
    """Función principal que orquesta todo el proceso"""
    try:
        crear_carpeta_logs()
        escribir_log("SISTEMA", "INICIO", "Inicio del proceso de extracción")
        
        # Descargar todos los excels (siempre sobrescribiendo)
        archivos_descargados = descargar_excels()
        if not archivos_descargados:
            raise Exception("No se pudo descargar ningún archivo Excel")
        
        # Cargar todos los excels a la memoria
        todos_los_excels_cargados = {}
        for nombre_archivo, ruta_archivo in tqdm(archivos_descargados.items(), desc="Cargando Excels en memoria"):
            try:
                todos_los_excels_cargados[nombre_archivo] = cargar_excel_completo(ruta_archivo)
            except Exception as e:
                error_msg = f"Error al cargar {nombre_archivo} en memoria: {str(e)}"
                print(f"\n{error_msg}")
                escribir_log("SISTEMA", "ERROR_CARGA", error_msg)
        
        if not todos_los_excels_cargados:
            raise Exception("No se pudo cargar ningún archivo Excel en memoria")

        # --- Creación de Índices ---
        todos_los_indices = {}
        for nombre_archivo, excel_data in tqdm(todos_los_excels_cargados.items(), desc="Creando índices de búsqueda"):
            todos_los_indices[nombre_archivo] = crear_indice_excel(excel_data)
        
        # Cargar archivo Excel con los códigos
        codigos_file = "Codigos.xlsx"
        try:
            wb = load_workbook(codigos_file)
            ws = wb.active
            # Estilos para marcado de errores
            red_font = Font(color="FF0000")
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            no_fill = PatternFill(fill_type=None)
            no_font = Font(color="000000")

            columnas_requeridas = ["Excel Origen", "ID", "Variable", "Pestaña Renombrada"]
            df_codigos = pd.read_excel(
                codigos_file,
                usecols=columnas_requeridas,
                dtype={col: str for col in columnas_requeridas} # Leer todo como texto
            )
            
            # Limpieza de datos de Codigos.xlsx
            df_codigos["Pestaña Renombrada"] = df_codigos["Pestaña Renombrada"].fillna("Otros")
            df_codigos["Pestaña Renombrada"] = df_codigos["Pestaña Renombrada"].replace({"nan": "Otros", "": "Otros"})
            
        except Exception as e:
            error_msg = f"Error al leer Codigos.xlsx: {str(e)}"
            print(error_msg)
            escribir_log("SISTEMA", "ERROR", error_msg)
            return

        # Cargar o inicializar BD
        try:
            bd_hojas = pd.read_excel("BD.xlsx", sheet_name=None)
            
            # Migrar datos de hojas 'nan' a 'Otros'
            for hoja in list(bd_hojas.keys()): # Usamos list() para evitar RuntimeError
                if pd.isna(hoja) or str(hoja).strip().lower() in ["nan", ""]:
                    datos = bd_hojas.pop(hoja)
                    if "Otros" not in bd_hojas:
                        bd_hojas["Otros"] = datos
                    else:
                        bd_hojas["Otros"] = pd.concat([bd_hojas["Otros"], datos])
            
            # Asegurar columna 'fecha' en todas las hojas
            for hoja in bd_hojas:
                if 'fecha' not in bd_hojas[hoja].columns:
                    bd_hojas[hoja]['fecha'] = pd.NaT
                # Convertir fecha a datetime para merges
                bd_hojas[hoja]['fecha'] = pd.to_datetime(bd_hojas[hoja]['fecha'])
                
        except FileNotFoundError:
            bd_hojas = {"Otros": pd.DataFrame(columns=['fecha'])}
        except Exception as e:
            error_msg = f"Error al leer BD.xlsx: {str(e)}"
            print(error_msg)
            escribir_log("SISTEMA", "ERROR", error_msg)
            bd_hojas = {"Otros": pd.DataFrame(columns=['fecha'])}

        total_series = len(df_codigos)
        series_exitosas = 0
        series_fallidas = 0
        filas_con_error = []
        filas_exitosas = [] # Para limpiar errores previos

        # --- Procesamiento principal (Iterando sobre Codigos.xlsx) ---
        print("\nProcesando series desde Codigos.xlsx...")
        for fila_idx, fila_codigo in tqdm(df_codigos.iterrows(), total=total_series, desc="Procesando IDs"):
            
            id_serie = str(fila_codigo["ID"]).strip()
            excel_origen = str(fila_codigo["Excel Origen"]).strip()
            variable_nombre = str(fila_codigo["Variable"]).strip()
            pestaña = str(fila_codigo["Pestaña Renombrada"]).strip()
            
            # Validaciones de la fila
            if not id_serie or id_serie.lower() == "nan":
                error_msg = "ID vacío"
                filas_con_error.append((fila_idx, error_msg))
                continue
            if not excel_origen or excel_origen.lower() == "nan":
                error_msg = f"Excel Origen vacío para ID {id_serie}"
                filas_con_error.append((fila_idx, error_msg))
                continue
            if not variable_nombre or variable_nombre.lower() == "nan":
                error_msg = f"Nombre de Variable vacío para ID {id_serie}"
                filas_con_error.append((fila_idx, error_msg))
                continue
            if not pestaña or pestaña.lower() == "nan":
                pestaña = "Otros" # Default a 'Otros'
            
            # Fila de excel 1-based + 1 de header
            fila_excel = fila_idx + 2 

            try:
                # 1. Encontrar el archivo Excel correcto usando el mapeo
                nombre_archivo = MAPEO_ORIGEN_ARCHIVO.get(excel_origen)
                if not nombre_archivo:
                    raise KeyError(f"Valor '{excel_origen}' no encontrado en MAPEO_ORIGEN_ARCHIVO")
                
                # 2. Obtener los datos del Excel y el Índice correspondientes
                excel_data = todos_los_excels_cargados.get(nombre_archivo)
                indice = todos_los_indices.get(nombre_archivo)
                if not excel_data or not indice:
                    raise FileNotFoundError(f"Archivo '{nombre_archivo}' no fue cargado o indexado")

                # 3. Buscar el ID en el índice (Búsqueda instantánea)
                ubicacion = indice.get(id_serie)
                if not ubicacion:
                    raise ValueError(f"ID '{id_serie}' no encontrado en el índice de '{nombre_archivo}'")

                # 4. Extraer la serie usando la ubicación
                categorias_api, valores_api = extraer_serie_desde_indice(excel_data, ubicacion)
                
                # Crear DataFrame con los datos
                df_api = pd.DataFrame({
                    'fecha': categorias_api,
                    variable_nombre: valores_api
                })
                # Eliminar filas donde la fecha no se pudo parsear
                df_api = df_api.dropna(subset=['fecha'])
                
                if df_api.empty:
                     raise ValueError(f"No se extrajeron datos válidos para el ID '{id_serie}'")

                # 5. Actualizar la BD
                if pestaña not in bd_hojas:
                    bd_hojas[pestaña] = pd.DataFrame(columns=['fecha'])
                    bd_hojas[pestaña]['fecha'] = pd.to_datetime(bd_hojas[pestaña]['fecha'])
                
                # Asegurar que la fecha de la hoja BD sea datetime
                bd_hojas[pestaña]['fecha'] = pd.to_datetime(bd_hojas[pestaña]['fecha'])

                # Lógica de actualización para preservar datos manuales antiguos
                
                df_bd_existente = bd_hojas[pestaña] # DataFrame de la hoja actual
                
                # 1. Verificar si la columna ya existe
                if variable_nombre not in df_bd_existente.columns:
                    # Caso 1: La serie es nueva. Simplemente hacer merge.
                    df_merged = pd.merge(
                        df_bd_existente,
                        df_api,
                        on='fecha',
                        how='outer'
                    )
                else:
                    # Caso 2: La serie ya existe. Preservar datos manuales.
                    
                    # 2a. Encontrar la fecha de corte (el primer día de la *nueva* data)
                    if df_api.empty or df_api['fecha'].isna().all():
                        df_merged = df_bd_existente
                    else:
                        primera_fecha_nueva = df_api['fecha'].min()
                        
                        # 2b. Separar los datos existentes en dos:
                        #    - Lo que queremos preservar (todo ANTES de la fecha de corte)
                        #    - Lo que vamos a sobreescribir (todo DESDE la fecha de corte)
                        df_preservar = df_bd_existente[df_bd_existente['fecha'] < primera_fecha_nueva].copy()
                        df_a_actualizar = df_bd_existente[df_bd_existente['fecha'] >= primera_fecha_nueva].copy()
                        
                        # 2c. De la parte a actualizar, nos quedamos con las *otras* columnas
                        # (eliminamos la data vieja de la serie que estamos actualizando)
                        if variable_nombre in df_a_actualizar.columns:
                            df_a_actualizar = df_a_actualizar.drop(columns=[variable_nombre])
                        
                        # 2d. Hacemos merge de las *otras columnas* con la *nueva data*
                        # Usamos 'outer' para asegurar que se incluyan todas las fechas
                        df_actualizado = pd.merge(
                            df_a_actualizar, # Base con otras series y fechas >= corte
                            df_api,          # Data nueva (con 'fecha' y 'variable_nombre')
                            on='fecha',
                            how='outer'
                        )
                        
                        # 2e. Juntamos lo preservado (datos manuales) con lo actualizado
                        df_merged = pd.concat([df_preservar, df_actualizado], ignore_index=True)

                # Ordenar por fecha y asegurarse que fecha es la primera columna
                df_merged = df_merged.sort_values('fecha').reset_index(drop=True)
                
                # Re-colocar 'fecha' al principio si se movió
                if 'fecha' in df_merged.columns:
                    cols = ['fecha'] + [col for col in df_merged.columns if col != 'fecha']
                    df_merged = df_merged[cols]

                bd_hojas[pestaña] = df_merged
                
                series_exitosas += 1
                filas_exitosas.append(fila_idx)
                escribir_log(id_serie, "OK", f"Encontrado en {nombre_archivo}. Registros: {len(df_api)}")
                
            except Exception as e:
                error_msg = str(e)
                # print(f"Error procesando ID {id_serie}: {error_msg}") # Descomentar para debug
                escribir_log(id_serie, "ERROR", error_msg)
                filas_con_error.append((fila_idx, error_msg))
                continue
                
        series_fallidas = len(filas_con_error)

        # --- BLOQUE DE GUARDADO FINAL CORREGIDO ---
        try:
            print("\nGenerando hoja de Referencias e Índices...")

            # 1. PREPARAR DATOS DE LA PRIMERA PESTAÑA (REFERENCIAS)
            lista_ref = []
            mapa_frecuencias = {
                'A': 'Anual', 'S': 'Semestral', 'T': 'Trimestral', 
                'M': 'Mensual', 'D': 'Diaria'
            }

            for _, fila in df_codigos.iterrows():
                pestana_renombrada = str(fila.get("Pestaña Renombrada", "Otros")).strip()
                origen_codigo = str(fila.get("Excel Origen", "")).strip()
                
                # A. Frecuencia
                letra_frec = pestana_renombrada[-1].upper() if pestana_renombrada else ""
                frecuencia = mapa_frecuencias.get(letra_frec, "Otra")
                
                # B. TEMA (Extraer texto antes de los dos puntos)
                tema = origen_codigo.split(':')[0].strip()
                
                # C. URL
                nombre_archivo_fisico = MAPEO_ORIGEN_ARCHIVO.get(origen_codigo)
                url_fuente = EXCEL_URLS.get(nombre_archivo_fisico, "URL no encontrada")
                
                lista_ref.append({
                    "Tema": tema,
                    "Variable": fila.get("Variable"),
                    "Frecuencia": frecuencia,
                    "ID": fila.get("ID"),
                    "Pestaña": pestana_renombrada,
                    "Fuente": url_fuente
                })

            # Crear y Ordenar DataFrame Referencias
            df_referencias = pd.DataFrame(lista_ref)
            columnas_ordenadas = ["Tema", "Variable", "Frecuencia", "ID", "Pestaña", "Fuente"]
            # Asegurarnos que existan todas las columnas (por si la lista vino vacía)
            for col in columnas_ordenadas:
                if col not in df_referencias.columns:
                    df_referencias[col] = ""
            
            df_referencias = df_referencias[columnas_ordenadas]
            df_referencias = df_referencias.sort_values(
                by=["Tema", "Fuente", "Pestaña", "Variable", "Frecuencia"]
            )

            # 2. GUARDAR TODO EN EL EXCEL (BD.xlsx)
            print("Guardando BD.xlsx...")
            with pd.ExcelWriter("BD.xlsx", engine='openpyxl', datetime_format='YYYY-MM-DD') as writer:
                
                # A. Escribir la hoja Referencias (NUEVA)
                df_referencias.to_excel(writer, sheet_name='Referencias', index=False)
                
                # B. Escribir el resto de las hojas de datos
                for hoja_original, datos in bd_hojas.items():
                    hoja_str = str(hoja_original).strip()[:31]
                    
                    # --- CORRECCIÓN CRÍTICA AQUÍ ---
                    # Si la hoja se llama 'Referencias', NO la guardamos en este ciclo
                    # (porque es la versión vieja cargada en memoria y pisaría la nueva)
                    if hoja_str == 'Referencias':
                        continue
                    # -------------------------------

                    columnas_de_datos = [col for col in datos.columns if col != 'fecha']
                    if datos.empty or not columnas_de_datos:
                        continue 

                    # Formateo de Fecha
                    frecuencia_hoja = hoja_str[-1].upper()
                    datos_a_guardar = datos.copy()
                    
                    if frecuencia_hoja in ['A', 'M', 'T']: 
                        datos_a_guardar['fecha'] = pd.to_datetime(datos_a_guardar['fecha']).dt.to_period('M')
                    else: 
                        datos_a_guardar['fecha'] = pd.to_datetime(datos_a_guardar['fecha']).dt.strftime('%Y-%m-%d')

                    datos_a_guardar.to_excel(writer, sheet_name=hoja_str, index=False)
            
            # --- FIN DE GUARDADO ---

            # Actualizar Codigos.xlsx con errores
            for fila_idx in filas_exitosas:
                fila_excel = fila_idx + 2
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=fila_excel, column=col).font = no_font
                    ws.cell(row=fila_excel, column=col).fill = no_fill
            
            for fila_idx, error_msg in filas_con_error:
                fila_excel = fila_idx + 2
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=fila_excel, column=col).font = red_font
            
            wb.save(codigos_file)
            
            if filas_con_error:
                print(f"\nSe marcaron en rojo {len(filas_con_error)} IDs con errores en {codigos_file}")
            
            resumen = f"Proceso completado. Series: {total_series} | Exitosas: {series_exitosas} | Fallidas: {series_fallidas}"
            print(f"\n{resumen}")
            escribir_log("SISTEMA", "FIN", resumen)
            
        except Exception as e:
            error_msg = f"Error al guardar archivos: {str(e)}"
            print(f"\n{error_msg}")
            escribir_log("SISTEMA", "ERROR", error_msg)

    except KeyboardInterrupt:
        print("\nProceso interrumpido por el usuario")
        escribir_log("SISTEMA", "INTERRUMPIDO", "Proceso detenido manualmente")
    except Exception as e:
        import traceback
        print(f"\nError inesperado: {str(e)}")
        print(traceback.format_exc()) # Imprimir stack trace para debug
        escribir_log("SISTEMA", "ERROR_CRITICO", str(e))

if __name__ == "__main__":
    procesar_datos()

